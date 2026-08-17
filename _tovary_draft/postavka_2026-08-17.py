#!/usr/bin/env python3
"""Сборка файлов для YClients: новые карточки товара + два прихода.
Поставка LUXIO/AKZENTZ/StraDerm от 17.08.2026, накладная 18 500 ₽."""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

OUT = "/tmp/claude-0/-home-user-claude-md/24f48d1b-45ae-5fc4-8d8f-a8aa6dbed170/scratchpad/out/"

HDR_FILL = PatternFill("solid", fgColor="F2F2F2")
HDR_FONT = Font(bold=True)


def sheet(wb, title, headers, rows, widths):
    ws = wb.active if wb.active.max_row == 1 and wb.active.max_column == 1 else wb.create_sheet()
    ws.title = title
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(1, c).font = HDR_FONT
        ws.cell(1, c).fill = HDR_FILL
        ws.cell(1, c).alignment = Alignment(wrap_text=True, vertical="center")
    for r in rows:
        ws.append(r)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    return ws


# ─── 1. НОВЫЕ ТОВАРЫ ──────────────────────────────────────────────────────────
# Колонки — как в «Шаблон новые товары в Yclients.xls»
NEW_HDR = ["ID категории", "Категория", "ID", "Название", "Название в чеке", "Артикул",
           "Штрихкод", "Цена продажи, ₽", "Себестоимость, ₽",
           "Единицы измерения для продажи", "Единицы измерения для списания",
           "Сколько единиц списания в единице продажи", "Масса Нетто, гр.",
           "Масса Брутто, гр.", "Критический остаток", "Желаемый остаток", "Комментарий"]

# (id_кат, категория, название, артикул, цена продажи, себестоимость)
NEW = [
    (1770042, "Luxio PRO", "289 - LUXIO GLOW 15 мл",   "LUX-GEL-289", 0, 1720),
    (1770042, "Luxio PRO", "290 - LUXIO SIZZLE 15 мл", "LUX-GEL-290", 0, 1720),
    (1770042, "Luxio PRO", "291 - LUXIO IGNITE 15 мл", "LUX-GEL-291", 0, 1720),
    (1770042, "Luxio PRO", "Гель для наращивания PRO FORMANCE - STRUCTURE LUXIO 7 гр", "ULC050", 0, 1740),
    (1770024, "Straderm PRO", "ReBuild Bloom многофункциональный лак Straderm 14 мл (PRO)", "C-114", 0, 1360),
    (1770096, "STRADERM", "ReBuild Natural многофункциональный лак Straderm 14 мл", "C-115", 2300, 1360),
]

rows = []
for cid, cat, title, art, price, cost in NEW:
    rows.append([cid, cat, "", title, title, art, "", price, cost,
                 "шт", "шт", 1, "", "", 0, 1, ""])
wb = Workbook()
sheet(wb, "Новые товары", NEW_HDR, rows,
      [13, 14, 6, 58, 58, 14, 14, 15, 16, 14, 14, 16, 13, 14, 13, 13, 14])
wb.save(OUT + "Новые товары в YClients — 17.08.2026.xlsx")

# ─── 2. ПРИХОДЫ ───────────────────────────────────────────────────────────────
# Колонки — как в «Шаблон по приходу товара на склад Yclients.xlsx»
PRIH_HDR = ["Название товара*", "Артикул ", "Штрихкод ", "Цена поставки*",
            "Количество*", "Скидка", "Итоговая цена"]

RABOTA = [   # склад расходников салона (637822)
    ("BUILD конструирующий гель LUXIO 15 ml", "LUX-GEL-010", 1860),
    ("289 - LUXIO GLOW 15 мл", "LUX-GEL-289", 1720),
    ("290 - LUXIO SIZZLE 15 мл", "LUX-GEL-290", 1720),
    ("291 - LUXIO IGNITE 15 мл", "LUX-GEL-291", 1720),
    ("Обезжириватель PREP & WIPE LUXIO 480 мл", "DS32", 3980),
    ("Гель для наращивания PRO FORMANCE - STRUCTURE LUXIO 7 гр", "ULC050", 1740),
    ("Гель для наращивания PRO FORMANCE - BALANCE Clear LUXIO 7 гр", "ULC060", 1740),
    ("ReBuild Bloom многофункциональный лак Straderm 14 мл (PRO)", "C-114", 1360),
]
PRODAZHA = [  # склад магазина (637823)
    ("KERAFIX восстанавливающий концентрат для ногтей Straderm 14 мл", "C-113", 1300),
    ("ReBuild Natural многофункциональный лак Straderm 14 мл", "C-115", 1360),
]


def prihod(name, items, expect):
    wb = Workbook()
    rows, total = [], 0
    for title, art, price in items:
        rows.append([title, art, "", price, 1, "", price])
        total += price
    rows.append(["", "", "", "", "", "ИТОГО", total])
    ws = sheet(wb, "Приход", PRIH_HDR, rows, [64, 14, 14, 15, 13, 10, 15])
    ws.cell(ws.max_row, 6).font = HDR_FONT
    ws.cell(ws.max_row, 7).font = HDR_FONT
    wb.save(OUT + name)
    assert total == expect, "%s: %s ≠ %s" % (name, total, expect)
    return total


t1 = prihod("Приход В РАБОТУ — 17.08.2026.xlsx", RABOTA, 15840)
t2 = prihod("Приход В ПРОДАЖУ — 17.08.2026.xlsx", PRODAZHA, 2660)
print("в работу:", t1, "| в продажу:", t2, "| всего:", t1 + t2)
assert t1 + t2 == 18500
print("сумма сходится с накладной: 18 500 ₽")
